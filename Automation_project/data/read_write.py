import openpyxl

class ReadWrite:
    def __init__(self,file):
        self.file=file

    # excel读
    def read(self):
        wk=openpyxl.load_workbook(self.file)
        sheets=wk.sheetnames
        table=wk[sheets[0]]
        table=wk.active
        list1=[]
        nrows=table.max_row
        #ncols=table.max_column
        for i in range(2,nrows+1):
            username=table.cell(i,1).value
            password=table.cell(i,2).value
            group=(username,password)
            list1.append(group)
        wk.close()
        return list1

    # excel写操作
    def write(self,username,password):
        wk=openpyxl.load_workbook(self.file)
        sheets=wk.sheetnames
        table=wk[sheets[0]]
        wk.active
        nrows=table.max_row
        for i in range(nrows+1,nrows+2):
            table.cell(i+1, 1).value=username
            table.cell(i+1, 2).value=password
        wk.save(self.file)
        wk.close()